from django.shortcuts import render
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
from users.decorators import staff_required
from .models import AuditLog
from django.contrib.auth import get_user_model
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

User = get_user_model()

@staff_required
def audit_logs_view(request):
    """
    Display audit logs with filtering capabilities.
    """
    # Get filter parameters
    user_id = request.GET.get('user')
    action_query = request.GET.get('action', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Base queryset
    logs = AuditLog.objects.all().select_related('user')
    
    # Apply filters
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    if action_query:
        logs = logs.filter(action__icontains=action_query)
    
    if start_date:
        logs = logs.filter(timestamp__gte=start_date)
    
    if end_date:
        # Add one day to include the entire end date
        from datetime import datetime, timedelta
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        logs = logs.filter(timestamp__lt=end_datetime)
    
    # Pagination
    paginator = Paginator(logs, 50)  # 50 logs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all users for filter dropdown
    users = User.objects.all().order_by('username')
    
    context = {
        'page_obj': page_obj,
        'users': users,
        'filters': {
            'user_id': user_id,
            'action': action_query,
            'start_date': start_date,
            'end_date': end_date,
        }
    }
    return render(request, 'audit/audit_logs.html', context)

@staff_required
def export_audit_logs_csv(request):
    """
    Export audit logs to CSV with applied filters.
    """
    # Get the same filters as the list view
    user_id = request.GET.get('user')
    action_query = request.GET.get('action', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Apply filters
    logs = AuditLog.objects.all().select_related('user')
    
    if user_id:
        logs = logs.filter(user_id=user_id)
    if action_query:
        logs = logs.filter(action__icontains=action_query)
    if start_date:
        logs = logs.filter(timestamp__gte=start_date)
    if end_date:
        from datetime import timedelta
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        logs = logs.filter(timestamp__lt=end_datetime)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="audit_logs_{timestamp_str}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Log ID', 'User', 'Action', 'Timestamp', 'IP Address', 'Details'])
    
    for log in logs:
        writer.writerow([
            log.log_id,
            log.user.username,
            log.action,
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.ip_address,
            log.details or ''
        ])
    
    return response

@staff_required
def export_audit_logs_excel(request):
    """
    Export audit logs to Excel with applied filters.
    """
    # Get the same filters as the list view
    user_id = request.GET.get('user')
    action_query = request.GET.get('action', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Apply filters
    logs = AuditLog.objects.all().select_related('user')
    
    if user_id:
        logs = logs.filter(user_id=user_id)
    if action_query:
        logs = logs.filter(action__icontains=action_query)
    if start_date:
        logs = logs.filter(timestamp__gte=start_date)
    if end_date:
        from datetime import timedelta
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        logs = logs.filter(timestamp__lt=end_datetime)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Log ID', 'User', 'Action', 'Timestamp', 'IP Address', 'Details']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log.log_id)
        ws.cell(row=row_idx, column=2, value=log.user.username)
        ws.cell(row=row_idx, column=3, value=log.action)
        ws.cell(row=row_idx, column=4, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=5, value=log.ip_address)
        ws.cell(row=row_idx, column=6, value=log.details or '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response