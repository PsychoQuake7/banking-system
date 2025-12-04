from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from django.http import HttpResponse
from io import BytesIO
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill


def generate_loan_applications_pdf(applications):
    """Generate Loan Applications Report PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Kayamanan Banking System - Loan Applications Report ({timezone.now().date()})", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.25*inch))
    
    # Summary
    try:
        total_count = applications.count()
    except (AttributeError, TypeError):
        total_count = len(applications)
    
    summary = Paragraph(f"Total Applications: {total_count}", styles['Heading2'])
    elements.append(summary)
    elements.append(Spacer(1, 0.25*inch))
    
    # Applications Table
    if applications:
        data = [['App ID', 'Client', 'Amount', 'Purpose', 'Applied', 'Eligibility', 'Officer', 'Status']]
        for app in applications:
            data.append([
                str(app.application_id),
                f"{app.client.first_name} {app.client.last_name}",
                f"₱{app.loan_amount:,.2f}",
                app.purpose[:20] + '...' if len(app.purpose) > 20 else app.purpose,
                app.application_date.strftime('%Y-%m-%d'),
                f"{app.eligibility_score or 0}%",
                app.loan_officer.username if app.loan_officer else 'Unassigned',
                app.get_status_display()
            ])
            
        t = Table(data, colWidths=[0.7*inch, 1.5*inch, 1.1*inch, 1.3*inch, 0.9*inch, 0.8*inch, 1*inch, 0.9*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No loan applications found.", styles['Normal']))
        
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="loan_applications_report.pdf"'
    return response


def generate_loan_applications_excel(applications):
    """Generate Loan Applications Report Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Applications"
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    # Title
    ws['A1'] = f"Kayamanan Banking System - Loan Applications Report"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Generated: {timezone.now().date()}"
    
    # Headers
    headers = ['App ID', 'Client Name', 'Email', 'Amount', 'Purpose', 'Term (Months)', 'Applied Date', 'Eligibility Score', 'Officer', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Data
    for row_num, app in enumerate(applications, 5):
        ws.cell(row=row_num, column=1, value=app.application_id)
        ws.cell(row=row_num, column=2, value=f"{app.client.first_name} {app.client.last_name}")
        ws.cell(row=row_num, column=3, value=app.client.user.email)
        ws.cell(row=row_num, column=4, value=float(app.loan_amount)).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=app.purpose)
        ws.cell(row=row_num, column=6, value=app.term_months)
        ws.cell(row=row_num, column=7, value=app.application_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=8, value=app.eligibility_score or 0)
        ws.cell(row=row_num, column=9, value=app.loan_officer.username if app.loan_officer else 'Unassigned')
        ws.cell(row=row_num, column=10, value=app.get_status_display())
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="loan_applications_report.xlsx"'
    return response
