from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from django.http import HttpResponse
from io import BytesIO
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill


def generate_clients_pdf(clients):
    """Generate Clients Report PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Kayamanan Banking System - Client Report ({timezone.now().date()})", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.25*inch))
    
    # Summary
    try:
        total_count = clients.count()
    except (AttributeError, TypeError):
        total_count = len(clients)
    
    summary = Paragraph(f"Total Clients: {total_count}", styles['Heading2'])
    elements.append(summary)
    elements.append(Spacer(1, 0.25*inch))
    
    # Clients Table
    if clients:
        data = [['Client ID', 'Name', 'Email', 'Credit Score', 'Monthly Income', 'Status']]
        for client in clients:
            data.append([
                str(client.client_id),
                f"{client.first_name} {client.last_name}",
                client.user.email,
                str(client.credit_score),
                f"₱{client.monthly_income:,.2f}",
                'Active' if client.user.is_active else 'Inactive'
            ])
            
        t = Table(data, colWidths=[1*inch, 2*inch, 2.5*inch, 1.2*inch, 1.5*inch, 1*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No clients found.", styles['Normal']))
        
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="clients_report.pdf"'
    return response


def generate_clients_excel(clients):
    """Generate Clients Report Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    # Title
    ws['A1'] = f"Kayamanan Banking System - Client Report"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Generated: {timezone.now().date()}"
    
    # Headers
    headers = ['Client ID', 'First Name', 'Last Name', 'Email', 'Phone', 'Credit Score', 'Monthly Income', 'Status', 'Created']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Data
    for row_num, client in enumerate(clients, 5):
        ws.cell(row=row_num, column=1, value=client.client_id)
        ws.cell(row=row_num, column=2, value=client.first_name)
        ws.cell(row=row_num, column=3, value=client.last_name)
        ws.cell(row=row_num, column=4, value=client.user.email)
        ws.cell(row=row_num, column=5, value=client.user.phone or 'N/A')
        ws.cell(row=row_num, column=6, value=client.credit_score)
        ws.cell(row=row_num, column=7, value=float(client.monthly_income)).number_format = '#,##0.00'
        ws.cell(row=row_num, column=8, value='Active' if client.user.is_active else 'Inactive')
        ws.cell(row=row_num, column=9, value=client.created_at.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clients_report.xlsx"'
    return response
