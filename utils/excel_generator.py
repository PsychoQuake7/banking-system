import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from django.utils import timezone

def generate_financial_report_excel(report_data, start_date, end_date):
    """
    Generate Financial Report Excel using openpyxl.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    
    # Styles
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    bold_font = Font(bold=True)
    
    # Title
    ws['A1'] = "Kayamanan Banking System - Income Statement"
    ws['A1'].font = title_font
    
    # Date Range
    ws['A2'] = f"Period: {start_date or 'Beginning'} to {end_date or 'Today'}"
    
    row_num = 4
    
    # Revenues
    ws[f'A{row_num}'] = "Revenues"
    ws[f'A{row_num}'].font = Font(size=14, bold=True)
    row_num += 1
    
    headers = ['Account Name', 'Code', 'Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    row_num += 1
    
    if report_data['revenues']:
        for acc in report_data['revenues']:
            ws.cell(row=row_num, column=1, value=acc['name'])
            ws.cell(row=row_num, column=2, value=acc['code'])
            ws.cell(row=row_num, column=3, value=acc['balance']).number_format = '#,##0.00'
            row_num += 1
        
        # Total Revenue
        ws.cell(row=row_num, column=1, value="Total Revenue").font = bold_font
        ws.cell(row=row_num, column=3, value=report_data['total_revenue']).font = bold_font
        ws.cell(row=row_num, column=3).number_format = '#,##0.00'
        row_num += 2
    else:
        ws.cell(row=row_num, column=1, value="No revenue recorded.")
        row_num += 2

    # Expenses
    ws[f'A{row_num}'] = "Expenses"
    ws[f'A{row_num}'].font = Font(size=14, bold=True)
    row_num += 1
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    row_num += 1
    
    if report_data['expenses']:
        for acc in report_data['expenses']:
            ws.cell(row=row_num, column=1, value=acc['name'])
            ws.cell(row=row_num, column=2, value=acc['code'])
            ws.cell(row=row_num, column=3, value=acc['balance']).number_format = '#,##0.00'
            row_num += 1
            
        # Total Expenses
        ws.cell(row=row_num, column=1, value="Total Expenses").font = bold_font
        ws.cell(row=row_num, column=3, value=report_data['total_expenses']).font = bold_font
        ws.cell(row=row_num, column=3).number_format = '#,##0.00'
        row_num += 2
    else:
        ws.cell(row=row_num, column=1, value="No expenses recorded.")
        row_num += 2
        
    # Net Income
    ws.cell(row=row_num, column=1, value="Net Income").font = Font(size=12, bold=True)
    cell = ws.cell(row=row_num, column=3, value=report_data['net_income'])
    cell.font = Font(size=12, bold=True, color='008000' if report_data['net_income'] >= 0 else 'FF0000')
    cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="financial_report.xlsx"'
    return response

def generate_loan_report_excel(active_loans, delinquent_loans, recent_repayments):
    """
    Generate Loan Reports Excel using openpyxl.
    """
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    # Helper to create sheet with headers
    def create_sheet(title, headers, data_rows):
        if title == "Active Loans":
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title=title)
            
        # Title
        ws['A1'] = f"{title} - {timezone.now().date()}"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            
        # Data
        for row_num, row_data in enumerate(data_rows, 4):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                # Format currency columns (heuristic: if 'Amount' or 'Balance' in header)
                if 'Amount' in headers[col_num-1] or 'Balance' in headers[col_num-1] or 'Overdue' in headers[col_num-1]:
                     if isinstance(cell_value, (int, float, type(None))): # Check type before formatting
                        cell.number_format = '#,##0.00'
        
        # Auto-adjust widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # 1. Active Loans
    active_data = []
    for loan in active_loans:
        active_data.append([
            loan.loan_id,
            f"{loan.application.client.first_name} {loan.application.client.last_name}",
            loan.start_date,
            loan.term_months,
            loan.loan_amount,
            loan.remaining_balance
        ])
    create_sheet("Active Loans", ['Loan ID', 'Client', 'Start Date', 'Term (Mos)', 'Amount', 'Balance'], active_data)
    
    # 2. Delinquent Accounts
    delinquent_data = []
    for item in delinquent_loans:
        delinquent_data.append([
            item['loan'].loan_id,
            f"{item['loan'].application.client.first_name} {item['loan'].application.client.last_name}",
            item['days_overdue'],
            item['total_overdue']
        ])
    create_sheet("Delinquent Accounts", ['Loan ID', 'Client', 'Days Overdue', 'Total Overdue'], delinquent_data)
    
    # 3. Recent Repayments
    repayment_data = []
    for trans in recent_repayments:
        loan_id = trans.loan.loan_id if trans.loan else '-'
        repayment_data.append([
            trans.transaction_date.date(),
            f"{trans.account.client.first_name} {trans.account.client.last_name}",
            loan_id,
            trans.amount
        ])
    create_sheet("Recent Repayments", ['Date', 'Client', 'Loan ID', 'Amount'], repayment_data)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="loan_reports.xlsx"'
    return response

def generate_balance_sheet_excel(report_data):
    """
    Generate Balance Sheet Excel using openpyxl.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    # Styles
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    bold_font = Font(bold=True)
    
    # Title
    ws['A1'] = "Kayamanan Banking System - Balance Sheet"
    ws['A1'].font = title_font
    
    # Date
    ws['A2'] = f"As of {report_data['as_of_date']}"
    
    row_num = 4
    
    # Helper for sections
    def write_section(title, accounts, total, total_label):
        nonlocal row_num
        ws[f'A{row_num}'] = title
        ws[f'A{row_num}'].font = Font(size=14, bold=True)
        row_num += 1
        
        headers = ['Account Name', 'Code', 'Amount']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        row_num += 1
        
        if accounts:
            for acc in accounts:
                ws.cell(row=row_num, column=1, value=acc['name'])
                ws.cell(row=row_num, column=2, value=acc['code'])
                ws.cell(row=row_num, column=3, value=acc['balance']).number_format = '#,##0.00'
                row_num += 1
            
            ws.cell(row=row_num, column=1, value=total_label).font = bold_font
            ws.cell(row=row_num, column=3, value=total).font = bold_font
            ws.cell(row=row_num, column=3).number_format = '#,##0.00'
            row_num += 2
        else:
            ws.cell(row=row_num, column=1, value=f"No {title.lower()} recorded.")
            row_num += 2

    write_section("Assets", report_data['assets'], report_data['total_assets'], "Total Assets")
    write_section("Liabilities", report_data['liabilities'], report_data['total_liabilities'], "Total Liabilities")
    write_section("Equity", report_data['equity'], report_data['total_equity'], "Total Equity")
    
    # Total Liabilities & Equity
    ws.cell(row=row_num, column=1, value="Total Liabilities & Equity").font = Font(size=12, bold=True)
    cell = ws.cell(row=row_num, column=3, value=report_data['total_liabilities_equity'])
    cell.font = Font(size=12, bold=True)
    cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="balance_sheet.xlsx"'
    return response
